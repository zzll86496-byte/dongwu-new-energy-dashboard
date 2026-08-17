from __future__ import annotations

import json
import math
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


WORKBOOK = Path(
    r"C:\Users\翟张墨涵\Desktop\【东吴电新】电动车数据库20260817\【东吴电新】电动车数据库20260817\常规类\销量\海外\【东吴电新】全球电动车销量数据库（更新至6月）20260726.xlsx"
)
FORECAST_WORKBOOK = Path(
    r"C:\Users\翟张墨涵\Desktop\【东吴电新】电动车数据库20260817\【东吴电新】电动车数据库20260817\预测类\【东吴电新】分区域2026年销量预期20260712.xlsx"
)
OUTPUT = Path(__file__).parents[1] / "app" / "global-sales" / "global-sales-data.json"
COUNTRY_INDEX_OUTPUT = Path(__file__).parents[1] / "app" / "global-sales" / "country-index.json"
COUNTRY_DATA_DIR = Path(__file__).parents[1] / "public" / "global-country-data"

MODULES = [
    ("region", "分区域", [("region", "区域结构", "分区域")]),
    ("europe", "欧洲分车型分车企", [("model", "分车型", "欧洲(分车型)"), ("company", "分车企", "欧洲(分车企)")]),
    ("usa", "美国分车型分车企", [("model", "分车型", "美国(分车型)"), ("company", "分车企", "美国(分车企)")]),
]

FORECAST_SHEETS = [
    ("china-caam", "中国中汽协", "中国中汽协"),
    ("china-local-pv", "中国本土乘用车", "中国本土乘用车"),
    ("china-export", "中国出口", "中国出口"),
    ("china-local-cv", "中国本土商用车", "中国本土商用车"),
    ("usa", "美国", "美国 "),
    ("europe", "欧洲", "欧洲"),
    ("other-countries", "其他国家", "其他国家"),
]


def period(value: Any) -> str | None:
    match = re.search(r"(20\d{2})(0[1-9]|1[0-2])", str(value or ""))
    return f"{match.group(1)}-{match.group(2)}" if match else None


def number(value: Any) -> float | int | None:
    if isinstance(value, bool) or not isinstance(value, (int, float)) or not math.isfinite(value):
        return None
    rounded = round(float(value), 4)
    return int(rounded) if rounded.is_integer() else rounded


def sparse(values: list[Any]) -> list[list[float | int]]:
    points: list[list[float | int]] = []
    for index, value in enumerate(values):
        numeric = number(value)
        if numeric is not None and numeric != 0:
            points.append([index, numeric])
    return points


def pivot_view(sheet: Any, key: str, title: str) -> dict[str, Any]:
    headers = list(next(sheet.iter_rows(min_row=6, max_row=6, values_only=True)))
    date_columns = [(index, period(value)) for index, value in enumerate(headers)]
    date_columns = [(index, value) for index, value in date_columns if value]
    periods = [value for _, value in date_columns]
    series = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
        name = str(row[1]).strip() if len(row) > 1 and row[1] not in (None, "") else ""
        if not name or name in {"总计", "Grand Total"}:
            continue
        values = [row[index] if index < len(row) else None for index, _ in date_columns]
        points = sparse(values)
        if not points:
            continue
        series.append({"id": f"{key}-{row_number}", "name": name, "points": points})
    return {
        "key": key,
        "title": title,
        "sourceSheet": sheet.title,
        "periods": periods,
        "series": series,
    }


def region_view(sheet: Any, key: str, title: str) -> dict[str, Any]:
    headers = list(next(sheet.iter_rows(min_row=2, max_row=2, values_only=True)))
    date_columns = [(index, period(value)) for index, value in enumerate(headers)]
    date_columns = [(index, value) for index, value in date_columns if value]
    periods = [value for _, value in date_columns]
    region = ""
    series = []
    # The lower half of this sheet is a separate penetration-rate table. This
    # page is a sales-volume database, so only use the first table (rows 3-47).
    for row_number, row in enumerate(sheet.iter_rows(min_row=3, max_row=47, values_only=True), start=3):
        if len(row) > 1 and row[1] not in (None, ""):
            region = str(row[1]).strip()
        kind = str(row[2]).strip() if len(row) > 2 and row[2] not in (None, "") else ""
        if not region or kind not in {"EV", "PHEV", "合计"}:
            continue
        values = [row[index] if index < len(row) else None for index, _ in date_columns]
        points = sparse(values)
        if not points:
            continue
        series.append({
            "id": f"{key}-{row_number}",
            "name": f"{region} · {kind}",
            "group": region,
            "kind": kind,
            "points": points,
        })
    return {
        "key": key,
        "title": title,
        "sourceSheet": sheet.title,
        "periods": periods,
        "series": series,
    }


def forecast_view(sheet: Any, key: str, title: str) -> dict[str, Any]:
    """Extract the first complete 2026 monthly scenario block.

    Several forecast sheets repeat the same data as raw vehicles, quarterly
    summaries and ten-thousand-vehicle tables. The first complete 12-month
    block is the source-granularity table; the compact CAAM sheet is already
    expressed in ten thousand vehicles and is normalised here to vehicles.
    """
    scenarios: dict[str, list[float | int]] = {}
    for row in range(1, sheet.max_row + 1):
        for column in range(1, sheet.max_column):
            if str(sheet.cell(row, column).value or "").strip() != "2026年":
                continue
            if str(sheet.cell(row, column + 1).value or "").strip() != "销量":
                continue
            monthly = [number(sheet.cell(row, column + offset).value) for offset in range(2, 14)]
            if any(value is None for value in monthly):
                continue
            scenario_text = str(sheet.cell(row + 1, column).value or "")
            scenario = next((name for name in ("乐观", "中性", "悲观") if name in scenario_text), None)
            if not scenario or scenario in scenarios:
                continue
            values = [float(value) for value in monthly if value is not None]
            # The compact CAAM table is in 万辆; other first-block tables are
            # raw vehicle counts. Annual magnitude makes the distinction safe.
            if sum(values) < 10_000:
                values = [value * 10_000 for value in values]
            scenarios[scenario] = [number(value) or 0 for value in values]
        if len(scenarios) == 3:
            break

    missing = [name for name in ("乐观", "中性", "悲观") if name not in scenarios]
    if missing:
        raise ValueError(f"{sheet.title}: missing 2026 monthly scenarios {missing}")

    periods = [f"2026-{month:02d}" for month in range(1, 13)]
    series = []
    for index, scenario in enumerate(("乐观", "中性", "悲观"), start=1):
        series.append({
            "id": f"forecast-{key}-{index}",
            "name": scenario,
            "kind": scenario,
            "points": sparse(scenarios[scenario]),
        })
    return {
        "key": key,
        "title": title,
        "sourceSheet": sheet.title.strip(),
        "periods": periods,
        "series": series,
    }


def build_country_data(workbook: Any) -> tuple[dict[str, Any], dict[str, Any]]:
    source_sheet = workbook["全球新能源车(原始数据)"]
    headers = list(next(source_sheet.iter_rows(min_row=2, max_row=2, values_only=True)))
    date_columns = [(index, period(value)) for index, value in enumerate(headers)]
    date_columns = [(index, value) for index, value in date_columns if value]
    periods = [value for _, value in date_columns]

    country_order: list[str] = []
    country_buckets: dict[str, dict[str, dict[str, list[float]]]] = {}
    invalid_names = {"", "-", "N/A", "n.a.", "NA", "未知"}
    country_aliases = {"阿拉伯酋长国": "阿联酋"}

    for row in source_sheet.iter_rows(min_row=3, values_only=True):
        country = country_aliases.get(str(row[0] or "").strip(), str(row[0] or "").strip())
        if not country:
            continue
        if country not in country_buckets:
            country_order.append(country)
            country_buckets[country] = {"model": {}, "company": {}}
        model = str(row[5] or "").strip()
        company = str((row[7] if len(row) > 7 and row[7] not in (None, "", "-") else row[1]) or "").strip()
        names = {"model": model, "company": company}
        numeric_values = [number(row[index] if index < len(row) else None) for index, _ in date_columns]
        if not any(value not in (None, 0) for value in numeric_values):
            continue
        for view_key, name in names.items():
            if name in invalid_names:
                continue
            values = country_buckets[country][view_key].setdefault(name, [0.0] * len(periods))
            for index, value in enumerate(numeric_values):
                if value not in (None, 0):
                    values[index] += float(value)

    penetration_sheet = workbook["新能源车渗透率(分国家)"]
    penetration_headers = list(next(penetration_sheet.iter_rows(min_row=3, max_row=3, values_only=True)))
    penetration_columns = [(index, period(value)) for index, value in enumerate(penetration_headers)]
    penetration_columns = [(index, value) for index, value in penetration_columns if value in set(periods)]
    penetration: dict[str, list[list[float | int]]] = {}
    for row in penetration_sheet.iter_rows(min_row=4, values_only=True):
        name = str(row[1] or "").strip() if len(row) > 1 else ""
        if not name or name in {"总计", "Grand Total"}:
            continue
        values = [row[index] if index < len(row) else None for index, _ in penetration_columns]
        penetration[name] = sparse(values)

    COUNTRY_DATA_DIR.mkdir(parents=True, exist_ok=True)
    for stale_file in COUNTRY_DATA_DIR.glob("country-*.json"):
        stale_file.unlink()
    index_entries = []
    for country_index, country in enumerate(country_order, start=1):
        key = f"country-{country_index:03d}"
        views = []
        for view_key, title in (("model", "分车型"), ("company", "分车企")):
            ranked = sorted(
                country_buckets[country][view_key].items(),
                key=lambda item: sum(item[1][-12:]),
                reverse=True,
            )
            series = [
                {"id": f"{view_key}-{series_index}", "name": name, "points": sparse(values)}
                for series_index, (name, values) in enumerate(ranked, start=1)
            ]
            views.append({
                "key": view_key,
                "title": title,
                "sourceSheet": source_sheet.title,
                "periods": periods,
                "series": series,
            })
        penetration_points = penetration.get(country, [])
        payload = {
            "country": country,
            "periods": periods,
            "penetration": {"name": country, "points": penetration_points},
            "views": views,
        }
        filename = f"{key}.json"
        (COUNTRY_DATA_DIR / filename).write_text(
            json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8"
        )
        index_entries.append({
            "key": key,
            "name": country,
            "file": filename,
            "modelCount": len(views[0]["series"]),
            "companyCount": len(views[1]["series"]),
            "hasPenetration": bool(penetration_points),
        })

    country_module = {
        "key": "countries",
        "title": "细分国家",
        "views": [
            {"key": "model", "title": "分车型", "sourceSheet": source_sheet.title, "periods": periods, "series": []},
            {"key": "company", "title": "分车企", "sourceSheet": source_sheet.title, "periods": periods, "series": []},
        ],
    }
    country_index_payload = {"updated": periods[-1], "countries": index_entries}
    return country_module, country_index_payload


def attach_module_penetration(workbook: Any, modules: list[dict[str, Any]]) -> None:
    """Attach the workbook's aggregate Europe and US penetration series."""
    sheet = workbook["新能源车渗透率(分国家)"]
    headers = list(next(sheet.iter_rows(min_row=3, max_row=3, values_only=True)))
    date_columns = [(index, period(value)) for index, value in enumerate(headers)]
    date_columns = [(index, value) for index, value in date_columns if value]
    periods = [value for _, value in date_columns]
    penetration: dict[str, list[list[float | int]]] = {}
    for row in sheet.iter_rows(min_row=4, values_only=True):
        name = str(row[1] or "").strip() if len(row) > 1 else ""
        if not name or name in {"总计", "Grand Total"}:
            continue
        values = [row[index] if index < len(row) else None for index, _ in date_columns]
        penetration[name] = sparse(values)

    labels = {"europe": "欧洲", "usa": "美国"}
    for module in modules:
        label = labels.get(module["key"])
        if not label:
            continue
        module["penetration"] = {
            "name": label,
            "sourceSheet": sheet.title,
            "periods": periods,
            "points": penetration.get(label, []),
        }


def main() -> None:
    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    modules = []
    for module_key, module_title, views in MODULES:
        extracted_views = []
        for view_key, view_title, sheet_name in views:
            sheet = workbook[sheet_name]
            extracted_views.append(
                region_view(sheet, view_key, view_title)
                if sheet_name == "分区域"
                else pivot_view(sheet, view_key, view_title)
            )
        modules.append({"key": module_key, "title": module_title, "views": extracted_views})

    attach_module_penetration(workbook, modules)

    country_module, country_index_payload = build_country_data(workbook)
    modules.append(country_module)
    COUNTRY_INDEX_OUTPUT.write_text(
        json.dumps(country_index_payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8"
    )

    forecast_workbook = load_workbook(FORECAST_WORKBOOK, read_only=False, data_only=True)
    forecast_views = [
        forecast_view(forecast_workbook[sheet_name], view_key, view_title)
        for view_key, view_title, sheet_name in FORECAST_SHEETS
    ]
    modules.append({"key": "forecast", "title": "分区域销量预期", "views": forecast_views})

    payload = {
        "updated": "2026-06",
        "source": WORKBOOK.name,
        "forecastSource": FORECAST_WORKBOOK.name,
        "modules": modules,
    }
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"wrote={OUTPUT}")
    print(f"bytes={OUTPUT.stat().st_size}")
    for module in modules:
        summary = ", ".join(f"{view['title']} {len(view['series'])}项/{len(view['periods'])}期" for view in module["views"])
        print(f"{module['title']}: {summary}")
    print(f"细分国家: {len(country_index_payload['countries'])}个国家")


if __name__ == "__main__":
    main()
